import openpyxl as xl


# this will add username and password in the xlsx(Excel) sheet:
def add_to_sheet(user_name,generated_password):

    wb = xl.load_workbook('pythonpasswords.xlsx')
    sheet = wb['Sheet1']
    empty_row_number = sheet.max_row+1

    # user_name cell:
    empty_row_cell = sheet.cell(empty_row_number,1)

    # password cell:
    empty_col_cell = sheet.cell(empty_row_number,2)

    empty_row_cell.value = user_name
    empty_col_cell.value = generated_password

    wb.save('pythonpasswords.xlsx')


# this fuction is used to show the password to the user who's user_name is Entered
def show_password(user_name):

    wb = xl.load_workbook('pythonpasswords.xlsx')
    sheet = wb['Sheet1']
    max_row= sheet.max_row
    
    i = 2
    
    while i<=max_row:
        current_user_name = sheet.cell(i,1).value
        if(current_user_name == user_name):
            required_password = sheet.cell(i,2).value
            return required_password
        i = i+1
        

    print(f"User_name : {user_name} not found! ")
    return None


# admin login code : this will display all the data in the sheet
def admin_login(admin_pass):
    if(admin_pass == ":::::"):
        print("user_names: ")
        wb = xl.load_workbook('pythonpasswords.xlsx')
        sheet = wb['Sheet1']
        
        i = 2
        while(i<=sheet.max_row):
            current_name_info = sheet.cell(i,1).value
            current_pass_info = sheet.cell(i,2).value

            print(f"{i-1}: user_name = {current_name_info} | password = {current_pass_info} ")
            i=i+1

    else:
        print("Wrong password !")